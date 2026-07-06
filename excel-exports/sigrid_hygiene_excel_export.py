#!/usr/bin/env python3

# Copyright Software Improvement Group
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import sys
from argparse import ArgumentParser
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

from objectives_excel_export import populateSystemDetailsSheet
from sigrid_api_client import SigridApiClient


def toExcel(activeSystems, metadata, objectives, users):
    workbook = Workbook()
    populateMetadataCompletenessSheet(workbook.create_sheet("Metadata completeness"), activeSystems, metadata)
    populateSnapshotFreshnessSheet(workbook.create_sheet("Snapshot freshness"), activeSystems, metadata)
    populateEolSystemsSheet(workbook.create_sheet("EOL systems"), metadata)
    populateLastSigridAccessSheet(workbook.create_sheet("Last Sigrid access"), users)
    populateSystemDetailsSheet(workbook.create_sheet("Objectives coverage"), activeSystems, metadata, objectives)
    del workbook["Sheet"]
    return workbook


def populateMetadataCompletenessSheet(sheet, activeSystems, metadata):
    sheet.append(["System name", "Division", "Team", "Supplier", "In production since", "Business criticality", "Lifecycle phase",
                  "Target industry", "Deployment type", "Application type", "Distribution strategy"])

    for system in activeSystems:
        systemName = metadata[system]["displayName"] or system
        division = "INCOMPLETE" if not metadata[system]["divisionName"] else metadata[system]["divisionName"]
        team = ", ".join(str(item) for item in (["INCOMPLETE"] if not metadata[system]["teamNames"] else metadata[system]["teamNames"]))
        supplier = "INCOMPLETE" if not metadata[system]["supplierNames"] else ""
        inProductionSince = "INCOMPLETE" if metadata[system]["inProductionSince"] is None else ""
        businessCriticality = "INCOMPLETE" if not metadata[system]["businessCriticality"] else ""
        lifecyclePhase = "INCOMPLETE" if not metadata[system]["lifecyclePhase"] else ""
        targetIndustry = "INCOMPLETE" if not metadata[system]["targetIndustry"] else ""
        deploymentType = "INCOMPLETE" if not metadata[system]["deploymentType"] else ""
        applicationType = "INCOMPLETE" if not metadata[system]["applicationType"] else ""
        distributionStrategy = "INCOMPLETE" if not metadata[system]["softwareDistributionStrategy"] else ""
        sheet.append([systemName, division, team, supplier, inProductionSince, businessCriticality, lifecyclePhase,
                      targetIndustry, deploymentType, applicationType, distributionStrategy])

    red_font = Font(color="FF0000")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "INCOMPLETE":
                cell.font = red_font


def populateSnapshotFreshnessSheet(sheet, activeSystems, metadata):
    sheet.append(["System name", "Division", "Team", "Snapshot freshness"])

    timeNow = datetime.now()
    for system in activeSystems:
        snapshotDate = sigrid.fetchArchitectureQuality(system)["snapshotDate"]
        if timeNow - datetime.fromisoformat(snapshotDate) > timedelta(days=90):
            systemName = metadata[system]["displayName"] or system
            division = "INCOMPLETE" if not metadata[system]["divisionName"] else metadata[system]["divisionName"]
            team = ", ".join(str(item) for item in (["INCOMPLETE"] if not metadata[system]["teamNames"] else metadata[system]["teamNames"]))
            sheet.append([systemName, division, team, ">3 months"])


def populateEolSystemsSheet(sheet, metadata):
    sheet.append(["System name", "Division", "Team", "Lifecycle phase", "Deactivated"])

    eolSystems = [name for name, meta in metadata.items() if meta["lifecyclePhase"] == "EOL"]
    activeSystems = [name for name, meta in metadata.items() if meta["active"] and not meta["isDevelopmentOnly"]]

    for system in eolSystems:
        systemName = metadata[system]["displayName"] or system
        division = "INCOMPLETE" if not metadata[system]["divisionName"] else metadata[system]["divisionName"]
        team = ", ".join(str(item) for item in (["INCOMPLETE"] if not metadata[system]["teamNames"] else metadata[system]["teamNames"]))
        deactivated = "yes" if system not in activeSystems else ""
        sheet.append([systemName, division, team, "Eol", deactivated])


def populateLastSigridAccessSheet(sheet, users):
    sheet.append(["Last name", "First name", "Email", "Role", "Last login"])

    timeNow = datetime.now()
    usersOlderThanOneYear = [user for user in users if user["lastLoginAt"] is not None and
                             timeNow - datetime.fromisoformat(user["lastLoginAt"]) > timedelta(days=365)]
    for user in usersOlderThanOneYear:
        sheet.append([user["lastName"], user["firstName"], user["email"], user["role"].title(), ">1 year"])


if __name__ == "__main__":
    parser = ArgumentParser(description="Excel export containing the Sigrid hygiene status of the portfolio.")
    parser.add_argument("--customer", type=str, required=True, help="Sigrid customer name.")
    parser.add_argument("--out", type=str, required=True, help="Output file.")
    parser.add_argument("--sigridurl", type=str, default="https://sigrid-says.com", help="Sigrid base URL.")
    args = parser.parse_args()

    if not os.environ.get("SIGRID_CI_TOKEN"):
        print("Missing Sigrid API token in environment variable SIGRID_CI_TOKEN")
        sys.exit(1)

    sigrid = SigridApiClient(args.sigridurl, args.customer)
    metadata = sigrid.fetchMetadata()
    activeSystems = [name for name, meta in metadata.items() if meta["active"] and not meta["isDevelopmentOnly"]]
    objectives = {eval["systemName"]: eval["objectives"] for eval in sigrid.fetchObjectivesEvaluation()}
    users = sigrid.fetchUsers()

    workbook = toExcel(activeSystems, metadata, objectives, users)
    workbook.save(os.path.expanduser(args.out))
