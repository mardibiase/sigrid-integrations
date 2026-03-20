#!/usr/bin/env python3

# Copyright Software Improvement Group
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import itertools
import os
import sys
from argparse import ArgumentParser
from collections import defaultdict
from openpyxl import Workbook

from sigrid_api_client import SigridApiClient


OBJECTIVE_DISPLAY_NAMES = {
    "OSH_MAX_SEVERITY" : "Open Source vulnerabilities",
    "OSH_MAX_FRESHNESS_RISK" : "Open Source freshness",
    "OSH_MAX_LICENSE_RISK" : "Open Source licenses",
    "SECURITY_MAX_SEVERITY" : "Security",
    "RELIABILITY_MAX_SEVERITY" : "Reliability"
}


def groupObjectivesByType(activeSystems, objectives):
    objectivesByType = defaultdict(list)
    for system in activeSystems:
        for objective in objectives[system]:
            objectivesByType[objective["type"]].append(objective)
    return objectivesByType


def formatObjectiveEvaluation(objectives, system, type):
    for objective in objectives[system]:
        if objective["type"] == type:
            if objective["targetMetAtEnd"] == "MET":
                return [formatTarget(objective), 1]
            elif objective["targetMetAtEnd"] == "NOT_MET":
                return [formatTarget(objective), 0]
    return ["", ""]


def formatObjectiveType(type):
    return OBJECTIVE_DISPLAY_NAMES.get(type, type.title().replace("_", " "))


def formatTarget(objective):
    return f"{objective['level'].title()} target: {objective['target']}"


def toExcel(activeSystems, metadata, objectives):
    workbook = Workbook()
    populatePerSystemSheet(workbook.create_sheet("Per system"), activeSystems, metadata, objectives)
    populatePerObjectiveSheet(workbook.create_sheet("Per objective"), activeSystems, objectives)
    populateSystemDetailsSheet(workbook.create_sheet("System details"), activeSystems, metadata, objectives)
    del workbook["Sheet"]
    return workbook


def populatePerSystemSheet(sheet, activeSystems, metadata, objectives):
    sheet.append(["System name", "Lifecyle phase", "Count of objectives met", "Count of objectives not met"])
    for system in activeSystems:
        displayName = metadata[system]["displayName"] or system
        lifecycle = (metadata[system]["lifecyclePhase"] or "").title()
        met = sum(1 for objective in objectives[system] if objective["targetMetAtEnd"] == "MET")
        unmet = sum(1 for objective in objectives[system] if objective["targetMetAtEnd"] == "NOT_MET")
        sheet.append([displayName, lifecycle, met, unmet])


def populatePerObjectiveSheet(sheet, activeSystems, objectives):
    objectivesByType = groupObjectivesByType(activeSystems, objectives)

    sheet.append(["Objective", "Number of systems where it is met", "Number of systems where it is not met"])
    for type in sorted(objectivesByType.keys()):
        displayName = formatObjectiveType(type)
        met = sum(1 for objective in objectivesByType[type] if objective["targetMetAtEnd"] == "MET")
        unmet = sum(1 for objective in objectivesByType[type] if objective["targetMetAtEnd"] == "NOT_MET")
        sheet.append([displayName, met, unmet])


def populateSystemDetailsSheet(sheet, activeSystems, metadata, objectives):
    types = sorted(groupObjectivesByType(activeSystems, objectives).keys())
    columns = [[f"Associated {formatObjectiveType(type)} objective", f"{formatObjectiveType(type)} objective met?"] for type in types]

    sheet.append(["System name", "Division", "Team"] + list(itertools.chain(*columns)))
    for system in activeSystems:
        displayName = metadata[system]["displayName"] or system
        division = "INCOMPLETE" if not metadata[system]["divisionName"] else metadata[system]["divisionName"]
        team = ", ".join(str(item) for item in (["INCOMPLETE"] if not metadata[system]["teamNames"] else metadata[system]["teamNames"]))
        evaluations = [formatObjectiveEvaluation(objectives, system, type) for type in types]
        sheet.append([displayName, division, team] + list(itertools.chain(*evaluations)))


if __name__ == "__main__":
    parser = ArgumentParser(description="Excel export containing the status of all objectives for all systems.")
    parser.add_argument("--customer", type=str, required=True, help="Sigrid customer name.")
    parser.add_argument("--out", type=str, required=True, help="Output file.")
    parser.add_argument("--sigridurl", type=str, default="https://sigrid-says.com", help="Sigrid base URL.")
    args = parser.parse_args()

    if not os.environ.get("SIGRID_CI_TOKEN"):
        print("Missing Sigrid API token in environment variable SIGRID_CI_TOKEN")
        sys.exit(1)

    sigrid = SigridApiClient(args.sigridurl, args.customer)
    metadata = sigrid.fetchMetadata()
    activeSystems = [name for name, meta in metadata.items() if meta["active"] and not meta["isDevelopmentOnly"]]
    objectives = {eval["systemName"]: eval["objectives"] for eval in sigrid.fetchObjectivesEvaluation()}

    workbook = toExcel(activeSystems, metadata, objectives)
    workbook.save(os.path.expanduser(args.out))
